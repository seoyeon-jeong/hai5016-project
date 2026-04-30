"""
scrape.py – University cafeteria menu scraper
Reads campus restaurant URLs from Excel, downloads each page, and uses an AI
model to extract structured menu items, saving them as JSON Lines.
"""

import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

import httpx
import openpyxl
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from loguru import logger
from openai import OpenAI

# ---------------------------------------------------------------------------
# Bootstrap
# ---------------------------------------------------------------------------

load_dotenv()

_now = datetime.now()
TODAY = _now.strftime("%Y-%m-%d")
SCRAPE_DATETIME = _now.strftime("%Y-%m-%d %H:%M:%S")

results_dir = Path("results")
logs_dir = Path("logs")
results_dir.mkdir(exist_ok=True)
logs_dir.mkdir(exist_ok=True)

OUTPUT_PATH = results_dir / f"menus-{TODAY}.jsonl"

# Configure loguru: INFO+ to console, DEBUG+ to file
logger.remove()
logger.add(
    sys.stderr,
    level="INFO",
    format="<green>{time:HH:mm:ss}</green> | <level>{level:<8}</level> | {message}",
    colorize=True,
)
logger.add(
    logs_dir / f"scrape-{TODAY}.log",
    level="DEBUG",
    format="{time:YYYY-MM-DD HH:mm:ss} | {level:<8} | {message}",
    encoding="utf-8",
    rotation="1 day",
)

# ---------------------------------------------------------------------------
# OpenAI client (Azure AI Foundry endpoint)
# ---------------------------------------------------------------------------

_client = OpenAI(
    api_key=os.environ["AZURE_FOUNDRY_API_KEY"],
    base_url=os.environ["AZURE_FOUNDRY_ENDPOINT"],
)
MODEL = os.environ["AZURE_FOUNDRY_MODEL"]

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8",
}

VALID_MEAL_TYPES = {"breakfast", "lunch", "dinner", "unknown"}

EXTRACTION_PROMPT = """\
You are extracting university cafeteria menu data from a webpage.

URL: {url}
University: {university}
Today's date: {scrape_date}

Page text (truncated):
{text}

Task: find every menu item listed and return a JSON array.
Each element must be an object with exactly these fields:
  "menu_date"  – date of the menu in YYYY-MM-DD format; use "{scrape_date}" if not stated.
  "meal_type"  – one of: "breakfast", "lunch", "dinner", "unknown".
  "meal_name"  – the name of a specific dish or food item (Korean or English).
               Only include if you are CONFIDENT it is a real food/dish name
               (e.g. "된장찌개", "비빔밥", "김치볶음밥").
               Do NOT include: section headers, table labels, calorie counts,
               column names, navigation text, or any non-food text.
               If you are unsure whether a string is a meal name, skip it.
  "price_krw"  – price in Korean Won as a digits-only string (e.g. "5000"),
               or "" if the price is not shown.
  "restaurant" – name of the cafeteria or dining hall as shown on the page,
               or "" if not found.

Rules:
- Return ONLY a valid JSON array. No markdown fences, no prose.
- If there are no valid menu items, return: []
"""


# ---------------------------------------------------------------------------
# Excel reader
# ---------------------------------------------------------------------------

def _infer_university(url: str) -> str:
    if "snu.ac.kr" in url:
        return "SNU"
    if "korea.ac.kr" in url:
        return "Korea University"
    if "skku.edu" in url:
        return "SKKU"
    return ""


def read_sites() -> list[dict]:
    """Return a list of {university, restaurant, url} dicts from the Excel file."""
    wb = openpyxl.load_workbook("campus_restaurant_websites.xlsx")
    ws = wb.active

    # Locate the header row (first row containing both 'University' and 'Url')
    header_col: dict[str, int] = {}
    header_row_idx = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and "University" in row and "Url" in row:
            header_col = {v: j for j, v in enumerate(row) if v is not None}
            header_row_idx = i
            break

    if not header_row_idx:
        raise ValueError("Could not find header row with 'University' and 'Url' columns")

    sites: list[dict] = []
    last_university = ""

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        def _cell(col_name: str) -> str:
            idx = header_col.get(col_name)
            if idx is None or idx >= len(row):
                return ""
            return (row[idx] or "").strip() if row[idx] else ""

        raw_univ = _cell("University")
        restaurant = _cell("Restaurant")
        url = _cell("Url")

        if not url:
            continue

        if raw_univ:
            last_university = raw_univ

        # URL inference takes priority for rows with no explicit university
        university = raw_univ or _infer_university(url) or last_university

        sites.append({"university": university, "restaurant": restaurant, "url": url})
        logger.debug(f"Site loaded: [{university}] {url}")

    return sites


# ---------------------------------------------------------------------------
# HTTP fetch
# ---------------------------------------------------------------------------

def fetch_page(url: str) -> str | None:
    """Download a URL and return its text, or None on failure / timeout."""
    try:
        with httpx.Client(headers=HEADERS, timeout=5.0, follow_redirects=True) as http:
            resp = http.get(url)
            resp.raise_for_status()
            return resp.text
    except httpx.TimeoutException:
        logger.warning(f"Timeout (>5 s) fetching {url} – skipping")
    except httpx.HTTPStatusError as exc:
        logger.warning(f"HTTP {exc.response.status_code} for {url} – skipping")
    except Exception as exc:
        logger.warning(f"Request error for {url}: {exc} – skipping")
    return None


def html_to_text(html: str) -> str:
    """Strip HTML tags and return cleaned plain text (max 12 000 chars)."""
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript", "iframe"]):
        tag.decompose()
    raw = soup.get_text(separator="\n")
    lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]
    return "\n".join(lines)[:12_000]


# ---------------------------------------------------------------------------
# AI extraction
# ---------------------------------------------------------------------------

def extract_menu(url: str, university: str, text: str) -> list[dict]:
    """Ask the model to extract menu items from page text. Returns raw list."""
    prompt = EXTRACTION_PROMPT.format(
        url=url,
        university=university,
        scrape_date=TODAY,
        text=text,
    )
    try:
        resp = _client.chat.completions.create(
            model=MODEL,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = resp.choices[0].message.content.strip()
        # Strip accidental markdown code fences
        raw = re.sub(r"^```(?:json)?\s*", "", raw, flags=re.MULTILINE)
        raw = re.sub(r"\s*```\s*$", "", raw, flags=re.MULTILINE)
        items = json.loads(raw)
        return items if isinstance(items, list) else []
    except json.JSONDecodeError as exc:
        logger.error(f"JSON parse error from model response for {url}: {exc}")
    except Exception as exc:
        logger.error(f"Model extraction failed for {url}: {exc}")
    return []


# ---------------------------------------------------------------------------
# Validation / normalisation
# ---------------------------------------------------------------------------

def normalise(item: dict, url: str, university: str, excel_restaurant: str) -> dict | None:
    """Validate and normalise one raw model item. Returns None to discard."""
    meal_name = str(item.get("meal_name", "")).strip()
    if len(meal_name) < 2:
        return None

    meal_type = str(item.get("meal_type", "unknown")).lower().strip()
    if meal_type not in VALID_MEAL_TYPES:
        meal_type = "unknown"

    menu_date = str(item.get("menu_date", TODAY)).strip()
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", menu_date):
        menu_date = TODAY

    # Keep only digits for price; empty string if none found
    price_krw = re.sub(r"[^\d]", "", str(item.get("price_krw", "")))

    # Prefer AI-extracted restaurant name; fall back to Excel value
    restaurant = str(item.get("restaurant", "")).strip() or excel_restaurant

    return {
        "scrape_date": SCRAPE_DATETIME,
        "url": url,
        "menu_date": menu_date,
        "meal_type": meal_type,
        "meal_name": meal_name,
        "price_krw": price_krw,
        "university": university,
        "restaurant": restaurant,
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    logger.info(f"=== Scrape started: {SCRAPE_DATETIME} ===")
    logger.info(f"Output: {OUTPUT_PATH}")

    sites = read_sites()
    logger.info(f"Loaded {len(sites)} sites from Excel")

    total_written = 0

    with open(OUTPUT_PATH, "a", encoding="utf-8") as out:
        for idx, site in enumerate(sites, 1):
            url = site["url"]
            university = site["university"]
            excel_restaurant = site["restaurant"]

            logger.info(f"[{idx}/{len(sites)}] {university or '?'} | {url}")

            html = fetch_page(url)
            if html is None:
                continue  # already logged in fetch_page

            text = html_to_text(html)
            logger.debug(f"  Page text length: {len(text)} chars")

            items = extract_menu(url, university, text)
            logger.info(f"  Model returned {len(items)} raw item(s)")

            site_written = 0
            for item in items:
                record = normalise(item, url, university, excel_restaurant)
                if record is None:
                    logger.debug(f"  Discarded: {item.get('meal_name', '')!r}")
                    continue
                out.write(json.dumps(record, ensure_ascii=False) + "\n")
                out.flush()
                total_written += 1
                site_written += 1
                logger.info(
                    f"  + [{record['meal_type']}] {record['meal_name']}"
                    f"  {record['menu_date']}  ₩{record['price_krw'] or '?'}"
                    f"  ({record['restaurant'] or 'unknown restaurant'})"
                )

            logger.info(f"  Saved {site_written} item(s) from this site")

    logger.info(f"=== Done. {total_written} total item(s) saved to {OUTPUT_PATH} ===")


if __name__ == "__main__":
    main()
